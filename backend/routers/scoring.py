from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from .. import models
from ..core.security import check_is_commissioner, get_current_user
from ..database import get_db
from ..schemas.scoring import ScoringRule, ScoringRuleCreate, ScoringTemplate
from ..services.scoring_import_service import (
    ScoringImportError,
    parse_csv_rows_to_preview,
    parse_csv_rows_to_rules,
)
from ..services.scoring_service import (
    active_scoring_rules_for_league,
    calculate_player_week_points,
    calculate_points_for_stats,
    recalculate_league_week_scores,
    recalculate_matchup_scores,
)

router = APIRouter(prefix="/scoring", tags=["scoring"])


class ScoringRuleUpdateRequest(BaseModel):
    category: str | None = None
    event_name: str | None = None
    description: str | None = None
    range_min: float | None = None
    range_max: float | None = None
    point_value: float | None = None
    calculation_type: str | None = None
    applicable_positions: list[str] | None = None
    position_ids: list[int] | None = None
    season_year: int | None = None
    source: str | None = None
    is_active: bool | None = None


class ScoringRuleUpsertItem(ScoringRuleCreate):
    id: int | None = None


class ScoringRuleBatchUpsertRequest(BaseModel):
    rules: list[ScoringRuleUpsertItem] = Field(default_factory=list)
    replace_existing_for_season: bool = False
    season_year: int | None = None


class TemplateWithRulesCreateRequest(BaseModel):
    name: str
    description: str | None = None
    season_year: int | None = None
    source_platform: str = "custom"
    is_system_template: bool = False
    rules: list[ScoringRuleCreate] = Field(default_factory=list)


class TemplateApplyRequest(BaseModel):
    season_year: int | None = None
    deactivate_existing: bool = True


class TemplateImportRequest(BaseModel):
    template_name: str
    season_year: int | None = None
    source_platform: str = "imported"
    csv_content: str


class ScoringImportPreviewRequest(BaseModel):
    csv_content: str
    season_year: int | None = None
    source_platform: str = "imported"


class ScoringImportApplyRequest(BaseModel):
    csv_content: str
    season_year: int | None = None
    source_platform: str = "imported"
    replace_existing_for_season: bool = False


class PlayerPointsUploadSummary(BaseModel):
    season: int
    week: int
    source: str
    rows_received: int
    rows_applied: int
    rows_invalid: int
    rows_deleted: int
    inserted: int
    updated: int
    player_id_column: str
    points_column: str


class ScoringRuleProposalCreateRequest(BaseModel):
    title: str
    description: str | None = None
    proposed_change: dict[str, Any]
    season_year: int | None = None
    voting_deadline: datetime | None = None


class ScoringRuleVoteRequest(BaseModel):
    vote: str  # yes|no|abstain
    comment: str | None = None
    vote_weight: float = 1.0


class ScoringRuleProposalFinalizeRequest(BaseModel):
    status: str  # approved|rejected|cancelled


class RuleSetResponse(BaseModel):
    league_id: int
    season_year: int | None
    active_rule_count: int
    rules: list[ScoringRule]


class ScoringPlayerPreviewRequest(BaseModel):
    player_id: int | None = None
    position: str | None = None
    season: int | None = None
    week: int | None = None
    season_year: int | None = None
    stats: dict[str, Any] = Field(default_factory=dict)


class ScoringWeekRecalcRequest(BaseModel):
    season: int
    season_year: int | None = None


class ScoringMatchupRecalcRequest(BaseModel):
    season: int
    season_year: int | None = None


class DraftAnalyzerPreviewItem(BaseModel):
    player_id: int | None = None
    player_name: str | None = None
    position: str
    stats: dict[str, Any] = Field(default_factory=dict)


class DraftAnalyzerPreviewRequest(BaseModel):
    season_year: int | None = None
    players: list[DraftAnalyzerPreviewItem] = Field(default_factory=list)


def _league_id_or_400(user: models.User) -> int:
    if not user.league_id:
        raise HTTPException(status_code=400, detail="User is not associated with a league")
    return int(user.league_id)


def _append_change_log(
    db: Session,
    *,
    league_id: int,
    scoring_rule_id: int | None,
    season_year: int | None,
    change_type: str,
    changed_by_user_id: int | None,
    rationale: str | None,
    previous_value: dict[str, Any] | None,
    new_value: dict[str, Any] | None,
) -> None:
    db.add(
        models.ScoringRuleChangeLog(
            league_id=league_id,
            scoring_rule_id=scoring_rule_id,
            season_year=season_year,
            change_type=change_type,
            rationale=rationale,
            previous_value=previous_value,
            new_value=new_value,
            changed_by_user_id=changed_by_user_id,
        )
    )


def _rule_to_dict(rule: models.ScoringRule) -> dict[str, Any]:
    return {
        "id": rule.id,
        "league_id": rule.league_id,
        "season_year": rule.season_year,
        "category": rule.category,
        "event_name": rule.event_name,
        "description": rule.description,
        "range_min": float(rule.range_min),
        "range_max": float(rule.range_max),
        "point_value": float(rule.point_value),
        "calculation_type": rule.calculation_type,
        "applicable_positions": rule.applicable_positions or [],
        "position_ids": rule.position_ids or [],
        "source": rule.source,
        "is_active": bool(rule.is_active),
        "template_id": rule.template_id,
    }


def _parse_csv_rules(csv_content: str) -> list[ScoringRuleCreate]:
    try:
        return parse_csv_rows_to_rules(csv_content)
    except ScoringImportError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def _normalize_header(value: str) -> str:
    return "".join(ch for ch in str(value or "").lower().strip().replace(" ", "_") if ch.isalnum() or ch == "_")


def _coerce_upload_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _resolve_column_name(headers: list[str], preferred: str | None, aliases: list[str]) -> str:
    normalized = {_normalize_header(name): name for name in headers}
    if preferred:
        preferred_key = _normalize_header(preferred)
        if preferred_key in normalized:
            return normalized[preferred_key]
        raise HTTPException(
            status_code=400,
            detail=f"Column '{preferred}' not found in upload. Available columns: {headers}",
        )

    for alias in aliases:
        alias_key = _normalize_header(alias)
        if alias_key in normalized:
            return normalized[alias_key]

    raise HTTPException(
        status_code=400,
        detail=f"Could not auto-detect required column. Available columns: {headers}",
    )


def _parse_csv_upload(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV upload is missing a header row")
    return [
        {
            (key or "").strip(): _coerce_upload_value(value)
            for key, value in (row or {}).items()
        }
        for row in reader
    ]


def _parse_xlsx_upload(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx uploads") from exc

    workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise HTTPException(status_code=400, detail="XLSX upload is empty")

    headers = [_coerce_upload_value(cell) for cell in rows[0]]
    if not any(headers):
        raise HTTPException(status_code=400, detail="XLSX upload is missing header columns")

    records: list[dict[str, str]] = []
    for raw_row in rows[1:]:
        record: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = raw_row[idx] if idx < len(raw_row) else None
            record[header] = _coerce_upload_value(value)
        records.append(record)
    return records


def _extract_points_override_rows(
    records: list[dict[str, str]],
    *,
    player_id_column: str | None,
    points_column: str | None,
) -> tuple[list[dict[str, Any]], str, str, int]:
    if not records:
        raise HTTPException(status_code=400, detail="Upload contains no data rows")

    headers = sorted({key for row in records for key in row.keys() if key})
    if not headers:
        raise HTTPException(status_code=400, detail="Upload contains no readable columns")

    resolved_player_col = _resolve_column_name(
        headers,
        player_id_column,
        aliases=["player_id", "player id", "playerid", "id"],
    )
    resolved_points_col = _resolve_column_name(
        headers,
        points_column,
        aliases=["fantasy_points", "fantasy points", "points", "score", "player_points"],
    )

    invalid_rows = 0
    parsed: list[dict[str, Any]] = []
    for row in records:
        raw_player = _coerce_upload_value(row.get(resolved_player_col))
        raw_points = _coerce_upload_value(row.get(resolved_points_col))
        if not raw_player and not raw_points:
            continue
        try:
            player_id = int(float(raw_player))
            fantasy_points = float(raw_points)
        except Exception:
            invalid_rows += 1
            continue
        parsed.append({"player_id": player_id, "fantasy_points": fantasy_points})

    if not parsed:
        raise HTTPException(status_code=400, detail="No valid Player ID/Points rows found in upload")

    return parsed, resolved_player_col, resolved_points_col, invalid_rows


def _apply_points_overrides(
    db: Session,
    *,
    season: int,
    week: int,
    source: str,
    rows: list[dict[str, Any]],
    replace_existing_for_source: bool,
) -> tuple[int, int, int, int]:
    # Last value wins when duplicate player IDs appear in the same upload.
    deduped = {int(item["player_id"]): float(item["fantasy_points"]) for item in rows}
    player_ids = list(deduped.keys())

    known_ids = {
        int(row[0])
        for row in db.query(models.Player.id).filter(models.Player.id.in_(player_ids)).all()
    }
    missing = [player_id for player_id in player_ids if player_id not in known_ids]
    if missing:
        sample = ", ".join(str(value) for value in missing[:10])
        raise HTTPException(status_code=400, detail=f"Unknown player_id values in upload: {sample}")

    rows_deleted = 0
    if replace_existing_for_source:
        rows_deleted = (
            db.query(models.PlayerWeeklyStat)
            .filter(
                models.PlayerWeeklyStat.season == season,
                models.PlayerWeeklyStat.week == week,
                models.PlayerWeeklyStat.source == source,
            )
            .delete(synchronize_session=False)
        )

    existing = {
        int(stat.player_id): stat
        for stat in (
            db.query(models.PlayerWeeklyStat)
            .filter(
                models.PlayerWeeklyStat.season == season,
                models.PlayerWeeklyStat.week == week,
                models.PlayerWeeklyStat.source == source,
                models.PlayerWeeklyStat.player_id.in_(player_ids),
            )
            .all()
        )
    }

    inserted = 0
    updated = 0
    for player_id, points in deduped.items():
        stat = existing.get(player_id)
        if stat:
            stat.fantasy_points = points
            updated += 1
            continue
        db.add(
            models.PlayerWeeklyStat(
                player_id=player_id,
                season=season,
                week=week,
                fantasy_points=points,
                stats={"imported_points_override": True},
                source=source,
            )
        )
        inserted += 1

    return inserted, updated, rows_deleted, len(deduped)


@router.get("/rules", response_model=list[ScoringRule])
def read_scoring_rules(
    season_year: int | None = Query(default=None),
    include_inactive: bool = Query(default=False),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    league_id = _league_id_or_400(current_user)

    query = db.query(models.ScoringRule).filter(models.ScoringRule.league_id == league_id)
    if season_year is not None:
        query = query.filter(models.ScoringRule.season_year == season_year)
    if not include_inactive:
        query = query.filter(models.ScoringRule.is_active.is_(True))

    return query.order_by(models.ScoringRule.event_name.asc(), models.ScoringRule.id.asc()).all()


@router.get("/rulesets/current", response_model=RuleSetResponse)
def read_current_ruleset(
    season_year: int | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    league_id = _league_id_or_400(current_user)
    effective_year = season_year

    query = db.query(models.ScoringRule).filter(
        models.ScoringRule.league_id == league_id,
        models.ScoringRule.is_active.is_(True),
    )
    if effective_year is not None:
        query = query.filter(models.ScoringRule.season_year == effective_year)

    rules = query.order_by(models.ScoringRule.event_name.asc(), models.ScoringRule.id.asc()).all()

    return RuleSetResponse(
        league_id=league_id,
        season_year=effective_year,
        active_rule_count=len(rules),
        rules=[ScoringRule.model_validate(r) for r in rules],
    )


@router.post("/calculate/player-preview")
def calculate_scoring_player_preview(
    request: ScoringPlayerPreviewRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    league_id = _league_id_or_400(current_user)

    if request.stats:
        rules = active_scoring_rules_for_league(db, league_id=league_id, season_year=request.season_year)
        total, breakdown = calculate_points_for_stats(
            stats=request.stats,
            position=request.position or "ALL",
            rules=rules,
        )
        return {
            "league_id": league_id,
            "player_id": request.player_id,
            "season": request.season,
            "week": request.week,
            "position": request.position,
            "points": total,
            "breakdown": [item.__dict__ for item in breakdown],
            "rules_evaluated": len(rules),
        }

    if request.player_id is None or request.season is None or request.week is None:
        raise HTTPException(
            status_code=400,
            detail="Provide stats payload or (player_id, season, week) for preview calculation",
        )

    points, breakdown, stats_payload = calculate_player_week_points(
        db,
        league_id=league_id,
        player_id=request.player_id,
        season=request.season,
        week=request.week,
        position=request.position,
        season_year=request.season_year,
    )

    return {
        "league_id": league_id,
        "player_id": request.player_id,
        "season": request.season,
        "week": request.week,
        "position": request.position,
        "points": points,
        "breakdown": [item.__dict__ for item in breakdown],
        "stats_used": stats_payload,
        "rules_evaluated": len(active_scoring_rules_for_league(db, league_id=league_id, season_year=request.season_year)),
    }


@router.post("/calculate/draft-analyzer-preview")
def calculate_draft_analyzer_preview(
    request: DraftAnalyzerPreviewRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    league_id = _league_id_or_400(current_user)
    rules = active_scoring_rules_for_league(db, league_id=league_id, season_year=request.season_year)

    results: list[dict[str, Any]] = []
    for item in request.players:
        points, breakdown = calculate_points_for_stats(
            stats=item.stats,
            position=item.position,
            rules=rules,
        )
        results.append(
            {
                "player_id": item.player_id,
                "player_name": item.player_name,
                "position": item.position,
                "projected_points": points,
                "breakdown": [row.__dict__ for row in breakdown],
            }
        )

    results.sort(key=lambda row: row["projected_points"], reverse=True)

    return {
        "league_id": league_id,
        "season_year": request.season_year,
        "rules_evaluated": len(rules),
        "players": results,
    }


@router.post("/import/preview")
def preview_scoring_import(
    request: ScoringImportPreviewRequest,
    current_user: models.User = Depends(check_is_commissioner),
):
    _league_id_or_400(current_user)

    try:
        preview = parse_csv_rows_to_preview(
            request.csv_content,
            source_platform=request.source_platform,
            season_year=request.season_year,
        )
    except ScoringImportError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {
        "row_count": len(preview),
        "source_platform": request.source_platform,
        "season_year": request.season_year,
        "rules": preview,
    }


@router.post("/import/apply", response_model=list[ScoringRule])
def apply_scoring_import(
    request: ScoringImportApplyRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    try:
        parsed_rules = parse_csv_rows_to_rules(
            request.csv_content,
            source_platform=request.source_platform,
            season_year=request.season_year,
        )
    except ScoringImportError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    touched_rule_ids: set[int] = set()
    created_rules: list[models.ScoringRule] = []

    for item in parsed_rules:
        payload = item.model_dump()
        rule = models.ScoringRule(
            **payload,
            league_id=league_id,
            created_by_user_id=current_user.id,
            updated_by_user_id=current_user.id,
        )
        db.add(rule)
        db.flush()
        touched_rule_ids.add(rule.id)
        created_rules.append(rule)

        _append_change_log(
            db,
            league_id=league_id,
            scoring_rule_id=rule.id,
            season_year=rule.season_year,
            change_type="imported",
            changed_by_user_id=current_user.id,
            rationale="Rule imported via /scoring/import/apply",
            previous_value=None,
            new_value=_rule_to_dict(rule),
        )

    if request.replace_existing_for_season:
        stale_query = db.query(models.ScoringRule).filter(
            models.ScoringRule.league_id == league_id,
            models.ScoringRule.is_active.is_(True),
        )
        if request.season_year is not None:
            stale_query = stale_query.filter(models.ScoringRule.season_year == request.season_year)

        stale_rules = [row for row in stale_query.all() if row.id not in touched_rule_ids]
        for stale in stale_rules:
            previous = _rule_to_dict(stale)
            stale.is_active = False
            stale.deactivated_at = datetime.now(timezone.utc)
            stale.updated_by_user_id = current_user.id
            _append_change_log(
                db,
                league_id=league_id,
                scoring_rule_id=stale.id,
                season_year=stale.season_year,
                change_type="deleted",
                changed_by_user_id=current_user.id,
                rationale="Rule deactivated by /scoring/import/apply replacement",
                previous_value=previous,
                new_value=_rule_to_dict(stale),
            )

    db.commit()
    for row in created_rules:
        db.refresh(row)

    return created_rules


@router.post("/import/upload-points-override", response_model=PlayerPointsUploadSummary)
async def upload_points_override(
    file: UploadFile = File(...),
    season: int = Query(..., ge=2000, le=2100),
    week: int = Query(..., ge=1, le=18),
    source: str = Query(default="manual_override"),
    replace_existing_for_source: bool = Query(default=True),
    player_id_column: str | None = Query(default=None),
    points_column: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    filename = (file.filename or "").strip()
    if not filename:
        raise HTTPException(status_code=400, detail="Uploaded file must include a filename")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        records = _parse_csv_upload(content)
    elif lower_name.endswith(".xlsx"):
        records = _parse_xlsx_upload(content)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use .csv or .xlsx")

    parsed_rows, resolved_player_col, resolved_points_col, invalid_rows = _extract_points_override_rows(
        records,
        player_id_column=player_id_column,
        points_column=points_column,
    )

    inserted, updated, rows_deleted, rows_applied = _apply_points_overrides(
        db,
        season=season,
        week=week,
        source=source,
        rows=parsed_rows,
        replace_existing_for_source=replace_existing_for_source,
    )

    _append_change_log(
        db,
        league_id=league_id,
        scoring_rule_id=None,
        season_year=season,
        change_type="stats_override_imported",
        changed_by_user_id=current_user.id,
        rationale="Player points override imported from uploaded file",
        previous_value=None,
        new_value={
            "filename": filename,
            "season": season,
            "week": week,
            "source": source,
            "replace_existing_for_source": replace_existing_for_source,
            "rows_received": len(records),
            "rows_applied": rows_applied,
            "rows_invalid": invalid_rows,
            "rows_deleted": rows_deleted,
            "inserted": inserted,
            "updated": updated,
            "player_id_column": resolved_player_col,
            "points_column": resolved_points_col,
        },
    )
    db.commit()

    return PlayerPointsUploadSummary(
        season=season,
        week=week,
        source=source,
        rows_received=len(records),
        rows_applied=rows_applied,
        rows_invalid=invalid_rows,
        rows_deleted=rows_deleted,
        inserted=inserted,
        updated=updated,
        player_id_column=resolved_player_col,
        points_column=resolved_points_col,
    )


@router.post("/calculate/weeks/{week}/recalculate")
def recalculate_week_scores(
    week: int,
    request: ScoringWeekRecalcRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    recalculated = recalculate_league_week_scores(
        db,
        league_id=league_id,
        week=week,
        season=request.season,
        season_year=request.season_year,
    )
    db.commit()

    return {
        "league_id": league_id,
        "week": week,
        "season": request.season,
        "recalculated_matchups": len(recalculated),
        "results": recalculated,
    }


@router.post("/calculate/matchups/{matchup_id}/recalculate")
def recalculate_single_matchup_score(
    matchup_id: int,
    request: ScoringMatchupRecalcRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    matchup = (
        db.query(models.Matchup)
        .filter(
            models.Matchup.id == matchup_id,
            models.Matchup.league_id == league_id,
        )
        .first()
    )
    if not matchup:
        raise HTTPException(status_code=404, detail="Matchup not found")

    result = recalculate_matchup_scores(
        db,
        matchup=matchup,
        season=request.season,
        season_year=request.season_year,
    )
    db.commit()
    return result


@router.post("/rules", response_model=ScoringRule)
def create_scoring_rule(
    rule: ScoringRuleCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)
    payload = rule.model_dump()

    db_rule = models.ScoringRule(
        **payload,
        league_id=league_id,
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
    )
    db.add(db_rule)
    db.flush()

    _append_change_log(
        db,
        league_id=league_id,
        scoring_rule_id=db_rule.id,
        season_year=db_rule.season_year,
        change_type="created",
        changed_by_user_id=current_user.id,
        rationale="Rule created via /scoring/rules",
        previous_value=None,
        new_value=_rule_to_dict(db_rule),
    )

    db.commit()
    db.refresh(db_rule)
    return db_rule


@router.put("/rules/{rule_id}", response_model=ScoringRule)
def update_scoring_rule(
    rule_id: int,
    request: ScoringRuleUpdateRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    rule = (
        db.query(models.ScoringRule)
        .filter(
            models.ScoringRule.id == rule_id,
            models.ScoringRule.league_id == league_id,
        )
        .first()
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")

    previous = _rule_to_dict(rule)
    updates = request.model_dump(exclude_unset=True)
    for key, value in updates.items():
        setattr(rule, key, value)

    rule.updated_by_user_id = current_user.id

    _append_change_log(
        db,
        league_id=league_id,
        scoring_rule_id=rule.id,
        season_year=rule.season_year,
        change_type="updated",
        changed_by_user_id=current_user.id,
        rationale="Rule updated via /scoring/rules/{rule_id}",
        previous_value=previous,
        new_value=_rule_to_dict(rule),
    )

    db.commit()
    db.refresh(rule)
    return rule


@router.delete("/rules/{rule_id}")
def deactivate_scoring_rule(
    rule_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    rule = (
        db.query(models.ScoringRule)
        .filter(
            models.ScoringRule.id == rule_id,
            models.ScoringRule.league_id == league_id,
        )
        .first()
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")

    previous = _rule_to_dict(rule)
    rule.is_active = False
    rule.deactivated_at = datetime.now(timezone.utc)
    rule.updated_by_user_id = current_user.id

    _append_change_log(
        db,
        league_id=league_id,
        scoring_rule_id=rule.id,
        season_year=rule.season_year,
        change_type="deleted",
        changed_by_user_id=current_user.id,
        rationale="Rule deactivated via /scoring/rules/{rule_id}",
        previous_value=previous,
        new_value=_rule_to_dict(rule),
    )

    db.commit()
    return {"ok": True, "id": rule_id}


@router.post("/rules/batch-upsert", response_model=list[ScoringRule])
def batch_upsert_scoring_rules(
    request: ScoringRuleBatchUpsertRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)
    if not request.rules:
        raise HTTPException(status_code=400, detail="At least one rule is required")

    season_year = request.season_year
    touched_ids: set[int] = set()
    results: list[models.ScoringRule] = []

    try:
        for item in request.rules:
            data = item.model_dump(exclude={"id"})
            if season_year is not None:
                data["season_year"] = season_year

            if item.id:
                rule = (
                    db.query(models.ScoringRule)
                    .filter(
                        models.ScoringRule.id == item.id,
                        models.ScoringRule.league_id == league_id,
                    )
                    .first()
                )
                if not rule:
                    raise HTTPException(status_code=404, detail=f"Rule {item.id} not found")

                previous = _rule_to_dict(rule)
                for key, value in data.items():
                    setattr(rule, key, value)
                rule.updated_by_user_id = current_user.id
                touched_ids.add(rule.id)

                _append_change_log(
                    db,
                    league_id=league_id,
                    scoring_rule_id=rule.id,
                    season_year=rule.season_year,
                    change_type="updated",
                    changed_by_user_id=current_user.id,
                    rationale="Rule updated via batch-upsert",
                    previous_value=previous,
                    new_value=_rule_to_dict(rule),
                )
                results.append(rule)
                continue

            rule = models.ScoringRule(
                **data,
                league_id=league_id,
                created_by_user_id=current_user.id,
                updated_by_user_id=current_user.id,
            )
            db.add(rule)
            db.flush()
            touched_ids.add(rule.id)

            _append_change_log(
                db,
                league_id=league_id,
                scoring_rule_id=rule.id,
                season_year=rule.season_year,
                change_type="created",
                changed_by_user_id=current_user.id,
                rationale="Rule created via batch-upsert",
                previous_value=None,
                new_value=_rule_to_dict(rule),
            )
            results.append(rule)

        if request.replace_existing_for_season:
            query = db.query(models.ScoringRule).filter(
                models.ScoringRule.league_id == league_id,
                models.ScoringRule.is_active.is_(True),
            )
            if season_year is not None:
                query = query.filter(models.ScoringRule.season_year == season_year)

            stale_rules = [r for r in query.all() if r.id not in touched_ids]
            now = datetime.now(timezone.utc)
            for stale in stale_rules:
                previous = _rule_to_dict(stale)
                stale.is_active = False
                stale.deactivated_at = now
                stale.updated_by_user_id = current_user.id
                _append_change_log(
                    db,
                    league_id=league_id,
                    scoring_rule_id=stale.id,
                    season_year=stale.season_year,
                    change_type="deleted",
                    changed_by_user_id=current_user.id,
                    rationale="Rule deactivated by batch-upsert replacement",
                    previous_value=previous,
                    new_value=_rule_to_dict(stale),
                )

        db.commit()
        for row in results:
            db.refresh(row)
        return results
    except Exception:
        db.rollback()
        raise


@router.get("/templates", response_model=list[ScoringTemplate])
def list_scoring_templates(
    season_year: int | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    league_id = _league_id_or_400(current_user)
    query = db.query(models.ScoringTemplate).filter(models.ScoringTemplate.league_id == league_id)
    if season_year is not None:
        query = query.filter(models.ScoringTemplate.season_year == season_year)
    return query.order_by(models.ScoringTemplate.name.asc(), models.ScoringTemplate.id.asc()).all()


@router.post("/templates", response_model=ScoringTemplate)
def create_scoring_template(
    request: TemplateWithRulesCreateRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    template = models.ScoringTemplate(
        league_id=league_id,
        season_year=request.season_year,
        name=request.name,
        description=request.description,
        source_platform=request.source_platform,
        is_system_template=request.is_system_template,
        created_by_user_id=current_user.id,
    )
    db.add(template)
    db.flush()

    for idx, rule in enumerate(request.rules):
        payload = rule.model_dump()
        payload.setdefault("season_year", request.season_year)
        payload.setdefault("source", "template")

        row = models.ScoringRule(
            **payload,
            league_id=league_id,
            template_id=template.id,
            created_by_user_id=current_user.id,
            updated_by_user_id=current_user.id,
        )
        db.add(row)
        db.flush()

        db.add(
            models.ScoringTemplateRule(
                template_id=template.id,
                scoring_rule_id=row.id,
                rule_order=idx,
                included=True,
            )
        )

        _append_change_log(
            db,
            league_id=league_id,
            scoring_rule_id=row.id,
            season_year=row.season_year,
            change_type="template_applied",
            changed_by_user_id=current_user.id,
            rationale=f"Rule added to template {template.name}",
            previous_value=None,
            new_value=_rule_to_dict(row),
        )

    db.commit()
    db.refresh(template)
    return template


@router.post("/templates/import", response_model=ScoringTemplate)
def import_scoring_template_from_csv(
    request: TemplateImportRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    try:
        rules = parse_csv_rows_to_rules(
            request.csv_content,
            source_platform=request.source_platform,
            season_year=request.season_year,
        )
    except ScoringImportError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return create_scoring_template(
        TemplateWithRulesCreateRequest(
            name=request.template_name,
            description=f"Imported from CSV via /scoring/templates/import",
            season_year=request.season_year,
            source_platform=request.source_platform,
            rules=rules,
        ),
        db=db,
        current_user=current_user,
    )


@router.get("/templates/{template_id}/export")
def export_scoring_template_to_csv(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    league_id = _league_id_or_400(current_user)

    template = (
        db.query(models.ScoringTemplate)
        .filter(
            models.ScoringTemplate.id == template_id,
            models.ScoringTemplate.league_id == league_id,
        )
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    rules = (
        db.query(models.ScoringRule)
        .filter(
            models.ScoringRule.template_id == template_id,
            models.ScoringRule.league_id == league_id,
        )
        .order_by(models.ScoringRule.id.asc())
        .all()
    )

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            "category",
            "event_name",
            "description",
            "range_min",
            "range_max",
            "point_value",
            "calculation_type",
            "applicable_positions",
            "position_ids",
            "season_year",
            "source",
        ],
    )
    writer.writeheader()
    for rule in rules:
        writer.writerow(
            {
                "category": rule.category,
                "event_name": rule.event_name,
                "description": rule.description or "",
                "range_min": float(rule.range_min),
                "range_max": float(rule.range_max),
                "point_value": float(rule.point_value),
                "calculation_type": rule.calculation_type,
                "applicable_positions": "|".join(rule.applicable_positions or []),
                "position_ids": "|".join(str(x) for x in (rule.position_ids or [])),
                "season_year": rule.season_year or "",
                "source": rule.source,
            }
        )

    return {
        "template_id": template.id,
        "template_name": template.name,
        "filename": f"scoring_template_{template.id}.csv",
        "csv": output.getvalue(),
    }


@router.post("/templates/{template_id}/apply", response_model=list[ScoringRule])
def apply_template_to_active_ruleset(
    template_id: int,
    request: TemplateApplyRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    template = (
        db.query(models.ScoringTemplate)
        .filter(
            models.ScoringTemplate.id == template_id,
            models.ScoringTemplate.league_id == league_id,
        )
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    template_rules = (
        db.query(models.ScoringRule)
        .filter(
            models.ScoringRule.template_id == template_id,
            models.ScoringRule.league_id == league_id,
            models.ScoringRule.is_active.is_(True),
        )
        .order_by(models.ScoringRule.id.asc())
        .all()
    )
    if not template_rules:
        raise HTTPException(status_code=400, detail="Template has no active rules")

    effective_year = request.season_year if request.season_year is not None else template.season_year

    if request.deactivate_existing:
        existing_query = db.query(models.ScoringRule).filter(
            models.ScoringRule.league_id == league_id,
            models.ScoringRule.is_active.is_(True),
            models.ScoringRule.template_id.is_(None),
        )
        if effective_year is not None:
            existing_query = existing_query.filter(models.ScoringRule.season_year == effective_year)

        for row in existing_query.all():
            previous = _rule_to_dict(row)
            row.is_active = False
            row.deactivated_at = datetime.now(timezone.utc)
            row.updated_by_user_id = current_user.id
            _append_change_log(
                db,
                league_id=league_id,
                scoring_rule_id=row.id,
                season_year=row.season_year,
                change_type="deleted",
                changed_by_user_id=current_user.id,
                rationale=f"Rule deactivated before applying template {template.name}",
                previous_value=previous,
                new_value=_rule_to_dict(row),
            )

    created_rules: list[models.ScoringRule] = []
    for src in template_rules:
        row = models.ScoringRule(
            league_id=league_id,
            season_year=effective_year,
            category=src.category,
            event_name=src.event_name,
            description=src.description,
            range_min=src.range_min,
            range_max=src.range_max,
            point_value=src.point_value,
            calculation_type=src.calculation_type,
            applicable_positions=src.applicable_positions,
            position_ids=src.position_ids,
            source="template",
            template_id=template.id,
            created_by_user_id=current_user.id,
            updated_by_user_id=current_user.id,
            is_active=True,
        )
        db.add(row)
        db.flush()

        _append_change_log(
            db,
            league_id=league_id,
            scoring_rule_id=row.id,
            season_year=row.season_year,
            change_type="template_applied",
            changed_by_user_id=current_user.id,
            rationale=f"Template {template.name} applied",
            previous_value=None,
            new_value=_rule_to_dict(row),
        )
        created_rules.append(row)

    db.commit()
    for row in created_rules:
        db.refresh(row)
    return created_rules


@router.post("/proposals")
def create_scoring_rule_proposal(
    request: ScoringRuleProposalCreateRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    proposal = models.ScoringRuleProposal(
        league_id=league_id,
        season_year=request.season_year,
        title=request.title,
        description=request.description,
        proposed_change=request.proposed_change,
        status="open",
        proposed_by_user_id=current_user.id,
        voting_deadline=request.voting_deadline,
    )
    db.add(proposal)
    db.commit()
    db.refresh(proposal)

    return {"id": proposal.id, "status": proposal.status}


@router.get("/proposals")
def list_scoring_rule_proposals(
    season_year: int | None = Query(default=None),
    status: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    league_id = _league_id_or_400(current_user)

    query = db.query(models.ScoringRuleProposal).filter(models.ScoringRuleProposal.league_id == league_id)
    if season_year is not None:
        query = query.filter(models.ScoringRuleProposal.season_year == season_year)
    if status:
        query = query.filter(models.ScoringRuleProposal.status == status)

    rows = query.order_by(models.ScoringRuleProposal.created_at.desc(), models.ScoringRuleProposal.id.desc()).all()
    return [
        {
            "id": row.id,
            "season_year": row.season_year,
            "title": row.title,
            "description": row.description,
            "status": row.status,
            "voting_deadline": row.voting_deadline,
            "created_at": row.created_at,
            "proposed_by_user_id": row.proposed_by_user_id,
            "vote_count": len(row.votes),
        }
        for row in rows
    ]


@router.post("/proposals/{proposal_id}/vote")
def vote_on_scoring_rule_proposal(
    proposal_id: int,
    request: ScoringRuleVoteRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    league_id = _league_id_or_400(current_user)

    proposal = (
        db.query(models.ScoringRuleProposal)
        .filter(
            models.ScoringRuleProposal.id == proposal_id,
            models.ScoringRuleProposal.league_id == league_id,
        )
        .first()
    )
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    if proposal.status != "open":
        raise HTTPException(status_code=400, detail="Proposal is not open for voting")

    vote_value = request.vote.lower().strip()
    if vote_value not in {"yes", "no", "abstain"}:
        raise HTTPException(status_code=400, detail="vote must be yes, no, or abstain")

    vote = (
        db.query(models.ScoringRuleVote)
        .filter(
            models.ScoringRuleVote.proposal_id == proposal_id,
            models.ScoringRuleVote.voter_user_id == current_user.id,
        )
        .first()
    )

    if vote:
        vote.vote = vote_value
        vote.comment = request.comment
        vote.vote_weight = request.vote_weight
        vote.voted_at = datetime.now(timezone.utc)
    else:
        vote = models.ScoringRuleVote(
            proposal_id=proposal_id,
            voter_user_id=current_user.id,
            vote=vote_value,
            comment=request.comment,
            vote_weight=request.vote_weight,
        )
        db.add(vote)

    db.commit()

    return {"proposal_id": proposal_id, "voter_user_id": current_user.id, "vote": vote_value}


@router.post("/proposals/{proposal_id}/finalize")
def finalize_scoring_rule_proposal(
    proposal_id: int,
    request: ScoringRuleProposalFinalizeRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_is_commissioner),
):
    league_id = _league_id_or_400(current_user)

    proposal = (
        db.query(models.ScoringRuleProposal)
        .filter(
            models.ScoringRuleProposal.id == proposal_id,
            models.ScoringRuleProposal.league_id == league_id,
        )
        .first()
    )
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")

    normalized_status = request.status.lower().strip()
    if normalized_status not in {"approved", "rejected", "cancelled"}:
        raise HTTPException(status_code=400, detail="status must be approved, rejected, or cancelled")

    proposal.status = normalized_status
    proposal.finalized_by_user_id = current_user.id
    proposal.finalized_at = datetime.now(timezone.utc)

    db.commit()

    return {"id": proposal.id, "status": proposal.status}


@router.get("/history")
def get_scoring_rule_history(
    season_year: int | None = Query(default=None),
    rule_id: int | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    league_id = _league_id_or_400(current_user)

    query = db.query(models.ScoringRuleChangeLog).filter(models.ScoringRuleChangeLog.league_id == league_id)
    if season_year is not None:
        query = query.filter(models.ScoringRuleChangeLog.season_year == season_year)
    if rule_id is not None:
        query = query.filter(models.ScoringRuleChangeLog.scoring_rule_id == rule_id)

    rows = (
        query.order_by(models.ScoringRuleChangeLog.changed_at.desc(), models.ScoringRuleChangeLog.id.desc())
        .limit(limit)
        .all()
    )

    return [
        {
            "id": row.id,
            "scoring_rule_id": row.scoring_rule_id,
            "season_year": row.season_year,
            "change_type": row.change_type,
            "rationale": row.rationale,
            "previous_value": row.previous_value,
            "new_value": row.new_value,
            "changed_by_user_id": row.changed_by_user_id,
            "changed_at": row.changed_at,
        }
        for row in rows
    ]
